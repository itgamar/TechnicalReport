import tkinter as tk
from openpyxl import load_workbook


class MainPage:

    def press_button_create(self):
        text_out = tk.Label(window, text="ff", fg="black", font=("Lucida Console", 12))
        text_out.grid(row=3, column=0, sticky="W")
        #wb = load_workbook('template.xlsx')
        ## grab the active worksheet
        #ws = wb.active
        ## Data can be assigned directly to cells
        #ws['C22'] = customer
        ##ws['V17'] = text_i_customer.get()
        ## Save the file
        #wb.save('prova.xlsx')

    def __init__(self, window):
        # insert window dimension here
        window.geometry("900x550")
        # ability to modify the size in terms of height and width
        window.resizable(True, True)
        window.configure(background="white")
        window.grid_columnconfigure(0, weight=1)
        # insert title here
        window.title("Rapporto tecnico")
        text_o_customer = "Cliente"
        #text_o_address = "Luogo intervento / indirizzo"
        # CUSTOMER
        text_o_customer = tk.Label(window, background="white", text=text_o_customer, fg="black", font=("Lucida Console", 12))
        text_o_customer.grid(row=1, column=0, sticky="W", padx=0, pady=0)
        text_i_customer = tk.Entry()
        text_i_customer.grid(row=2, column=0, sticky="WE", padx=10, pady=0)
        #text_i_customer.insert(0, "a default value")
        # ADDRESS
        #text_o_address = tk.Label(window, background="white", text=text_o_address, fg="black", font=("Lucida Console", 12))
        #text_o_address.grid(row=3, column=0, sticky="W", padx=0, pady=0)
        #text_i_address = tk.Entry()
        #text_i_address.grid(row=4, column=0, sticky="WE", padx=10, pady=0)
        # CREATE BUTTON
        button_create = tk.Button(text="Create xlsx", command=self.press_button_create())
        button_create.grid(row=100, column=0, padx=10, pady=20, sticky="W")



if __name__ == "__main__":
    window = tk.Tk()
    w = MainPage(window)
    window.mainloop()
    