from openpyxl import load_workbook

def press_button_create():
    wb = load_workbook('template.xlsx')

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['C22'] = text
    ws['V17'] = 'pippo'

    # Save the file
    wb.save('prova.xlsx')